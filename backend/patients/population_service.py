"""Aholi bazasi — qidiruv, sinxronlash, Excel import/export."""
from __future__ import annotations

import io
import re
from typing import Any

from django.db import transaction
from django.db.models import Q
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .models import Patient, PopulationRecord
from .passport_serial import normalize_passport_serial, validate_passport_serial_format
from .phone import normalize_patient_phone, patient_phone_variants
from .registry_number import registry_number_lookup_q

POPULATION_FIELDS = (
    'first_name', 'last_name', 'father_name', 'age', 'gender',
    'phone', 'address', 'region_id', 'district_id', 'anamnesis',
)

PRIMARY_CARE_POPULATION_FIELDS = (
    'birth_date', 'health_group', 'brigade',
    'risk_pregnant', 'risk_disabled', 'risk_chronic',
    'risk_social_vulnerable', 'risk_lone_elderly', 'risk_needs_care',
)

EXCEL_HEADERS = [
    '№', 'Исм', 'Фамилия', 'Отасини исми', 'Yosh', 'Jins',
    'Telefon', 'Pasport seriya raqami', 'Manzil', 'Anamnez vitae',
    'Tug\'ilgan sana', 'Sog\'liq guruhi', 'Brigada kodi',
    'Xavf: homilador', 'Xavf: nogiron', 'Xavf: surunkali',
    'Xavf: ijtimoiy', 'Xavf: yolg\'iz keksa', 'Xavf: parvarish',
]

_HEADER_ALIASES: dict[str, str] = {
    'no': 'row_num', 'n': 'row_num', 'tartib': 'row_num',
    'ism': 'first_name', 'имя': 'first_name', 'first_name': 'first_name', 'first name': 'first_name',
    'familiya': 'last_name', 'фамилия': 'last_name', 'last_name': 'last_name', 'last name': 'last_name',
    'otasini ismi': 'father_name', 'отасини исми': 'father_name', 'father_name': 'father_name',
    'otasining ismi': 'father_name', 'отечество': 'father_name',
    'yosh': 'age', 'возраст': 'age', 'age': 'age',
    'jins': 'gender', 'пол': 'gender', 'gender': 'gender',
    'telefon': 'phone', 'телефон': 'phone', 'phone': 'phone',
    'pasport seriya raqami': 'registry_number', 'pasport': 'registry_number',
    'паспорт': 'registry_number', 'registry_number': 'registry_number', 'passport': 'registry_number',
    'manzil': 'address', 'манзил': 'address', 'address': 'address',
    'anamnez vitae': 'anamnesis', 'anamnez': 'anamnesis', 'анамнез': 'anamnesis', 'anamnesis': 'anamnesis',
    'shikoyatlar': 'anamnesis', 'complaints': 'anamnesis',
    'tugilgan sana': 'birth_date', 'birth_date': 'birth_date', 'дата рождения': 'birth_date',
    'sog\'liq guruhi': 'health_group', 'health_group': 'health_group', 'группа здоровья': 'health_group',
    'brigada kodi': 'brigade_code', 'brigade_code': 'brigade_code', 'brigada': 'brigade_code',
    'xavf: homilador': 'risk_pregnant', 'risk_pregnant': 'risk_pregnant',
    'xavf: nogiron': 'risk_disabled', 'risk_disabled': 'risk_disabled',
    'xavf: surunkali': 'risk_chronic', 'risk_chronic': 'risk_chronic',
    'xavf: ijtimoiy': 'risk_social_vulnerable', 'risk_social_vulnerable': 'risk_social_vulnerable',
    'xavf: yolg\'iz keksa': 'risk_lone_elderly', 'risk_lone_elderly': 'risk_lone_elderly',
    'xavf: parvarish': 'risk_needs_care', 'risk_needs_care': 'risk_needs_care',
}


def _normalize_header(value: str) -> str:
    return re.sub(r'\s+', ' ', (value or '').strip().lower())


def _map_gender(raw: str) -> str:
    s = (raw or '').strip().lower()
    if s in ('erkak', 'm', 'male', 'м', 'муж', 'мужской', 'e'):
        return 'male'
    if s in ('ayol', 'f', 'female', 'ж', 'жен', 'женский', 'a'):
        return 'female'
    if s in ('boshqa', 'other', 'др', 'другое'):
        return 'other'
    return ''


def _region_names() -> tuple[dict[str, str], dict[str, dict]]:
    from .address_data import load_address_catalog
    catalog = load_address_catalog()
    region_names = {r['id']: r['name_uz'] for r in catalog['regions']}
    district_meta: dict[str, dict] = {}
    for r in catalog['regions']:
        for d in r.get('districts', []):
            district_meta[str(d['id'])] = {
                'district_name': d['name_uz'],
                'region_id': r['id'],
                'region_name': r['name_uz'],
            }
    return region_names, district_meta


def population_to_dict(rec: PopulationRecord) -> dict[str, Any]:
    region_names, district_meta = _region_names()
    meta = district_meta.get(str(rec.district_id or ''), {})
    return {
        'id': rec.id,
        'population_id': rec.id,
        'source': 'population',
        'registry_number': rec.registry_number,
        'first_name': rec.first_name,
        'last_name': rec.last_name,
        'father_name': rec.father_name or '',
        'age': rec.age or '',
        'gender': rec.gender or '',
        'phone': rec.phone or '',
        'address': rec.address or '',
        'region_id': rec.region_id or '',
        'district_id': rec.district_id or '',
        'region_name': region_names.get(str(rec.region_id or ''), meta.get('region_name', '')),
        'district_name': meta.get('district_name', ''),
        'anamnesis': rec.anamnesis or '',
        'birth_date': rec.birth_date.isoformat() if rec.birth_date else '',
        'health_group': rec.health_group or '',
        'brigade_id': rec.brigade_id,
        'brigade_name': rec.brigade.name if rec.brigade else '',
        'next_checkup_date': rec.next_checkup_date.isoformat() if rec.next_checkup_date else '',
        'last_checkup_date': rec.last_checkup_date.isoformat() if rec.last_checkup_date else '',
        'dispensary_registered': rec.dispensary_registered,
        'risk_pregnant': rec.risk_pregnant,
        'risk_disabled': rec.risk_disabled,
        'risk_chronic': rec.risk_chronic,
        'registered_by': str(getattr(rec.created_by, 'name', '') or '') if rec.created_by else '',
        'created_at': rec.created_at.isoformat() if rec.created_at else '',
        'updated_at': rec.updated_at.isoformat() if rec.updated_at else '',
        'analysis_count': 0,
        'last_analysis_at': '',
        'last_diagnosis': '',
        'last_complaint': rec.anamnesis or '',
        'last_physician': '',
        'can_view_clinical': False,
        'is_patient': False,
    }


def search_population(q: str, *, user=None, limit: int = 20) -> list[PopulationRecord]:
    q = (q or '').strip()
    if not q:
        return []
    qs = PopulationRecord.objects.all()
    if user:
        from .primary_care_access import population_for_user
        qs = population_for_user(user)
    lookup = registry_number_lookup_q(q)
    if lookup is not None:
        return list(qs.filter(lookup | Q(phone__icontains=q)).order_by('-updated_at')[:limit])
    tokens = [t for t in q.split() if t]
    clause = (
        Q(first_name__icontains=q)
        | Q(last_name__icontains=q)
        | Q(father_name__icontains=q)
        | Q(phone__icontains=q)
        | Q(registry_number__icontains=q)
        | Q(address__icontains=q)
    )
    if len(tokens) >= 2:
        clause |= Q(first_name__icontains=tokens[0], last_name__icontains=tokens[-1])
    elif len(tokens) == 1:
        token = tokens[0]
        clause |= Q(first_name__istartswith=token) | Q(last_name__istartswith=token)
    return list(qs.filter(clause).order_by('-updated_at')[:limit])


def find_population_by_registry(registry_number: str | None) -> PopulationRecord | None:
    rn = normalize_passport_serial(registry_number)
    if not rn:
        return None
    lookup = registry_number_lookup_q(rn)
    if lookup is not None:
        hit = PopulationRecord.objects.filter(lookup).order_by('-updated_at').first()
        if hit:
            return hit
    return PopulationRecord.objects.filter(registry_number__iexact=rn).first()


def _population_owned_by_user(user, rec: PopulationRecord) -> bool:
    if not user:
        return True
    from .primary_care_access import population_for_user, user_clinic_group_id
    if not user_clinic_group_id(user):
        return True
    return population_for_user(user).filter(pk=rec.pk).exists()


def apply_population_fields(rec: PopulationRecord, data: dict, *, user=None) -> PopulationRecord:
    changed = []
    all_fields = POPULATION_FIELDS + PRIMARY_CARE_POPULATION_FIELDS
    for field in all_fields:
        val = data.get(field)
        if val is None:
            continue
        if isinstance(val, str) and field not in PRIMARY_CARE_POPULATION_FIELDS:
            val = val.strip()
        if field == 'phone':
            val = normalize_patient_phone(val) or val
        if field == 'gender':
            val = _map_gender(val) if val and val not in ('male', 'female', 'other') else val
        if getattr(rec, field) != val:
            setattr(rec, field, val)
            changed.append(field)
    if user and changed:
        rec.updated_by = user
        changed.append('updated_by')
    if changed:
        rec.save(update_fields=changed + ['updated_at'])
    return rec


@transaction.atomic
def upsert_population_from_data(
    data: dict,
    *,
    user=None,
    source: str = 'manual',
    registry_number: str | None = None,
) -> PopulationRecord:
    rn = normalize_passport_serial(registry_number or data.get('registry_number'))
    if not rn:
        raise ValueError('Pasport seriya raqami kerak')
    existing = find_population_by_registry(rn)
    if existing and user and not _population_owned_by_user(user, existing):
        if source == 'patient_auto':
            # Shifokor bemor yaratganda — demografik ma'lumotni yangilash, ko'rinishni ta'minlash
            if not existing.created_by_id:
                existing.created_by = user
                existing.save(update_fields=['created_by', 'updated_at'])
            payload = {k: data.get(k) for k in POPULATION_FIELDS + PRIMARY_CARE_POPULATION_FIELDS if data.get(k) is not None}
            rec = apply_population_fields(existing, payload, user=user)
            from .primary_care_service import on_population_saved
            sync_meta = on_population_saved(rec, is_new=False)
            rec._primary_care_sync = sync_meta  # noqa: SLF001
            return rec
        raise ValueError('Boshqa klinika aholi yozuvini o\'zgartirish mumkin emas')
    payload = {k: data.get(k) for k in POPULATION_FIELDS + PRIMARY_CARE_POPULATION_FIELDS if data.get(k) is not None}
    if existing:
        rec = apply_population_fields(existing, payload, user=user)
        from .primary_care_service import on_population_saved
        sync_meta = on_population_saved(rec, is_new=False)
        rec._primary_care_sync = sync_meta  # noqa: SLF001
        return rec
    rec = PopulationRecord.objects.create(
        registry_number=rn,
        first_name=(data.get('first_name') or '').strip() or '—',
        last_name=(data.get('last_name') or '').strip() or '—',
        father_name=(data.get('father_name') or '').strip(),
        age=(data.get('age') or '').strip(),
        gender=_map_gender(data.get('gender', '')) if data.get('gender') else (data.get('gender') or ''),
        phone=normalize_patient_phone(data.get('phone')) or (data.get('phone') or '').strip(),
        address=(data.get('address') or '').strip(),
        region_id=(data.get('region_id') or '').strip(),
        district_id=(data.get('district_id') or '').strip(),
        anamnesis=(data.get('anamnesis') or data.get('complaints') or '').strip(),
        birth_date=data.get('birth_date') or None,
        health_group=(data.get('health_group') or '').strip(),
        brigade_id=data.get('brigade') or data.get('brigade_id') or None,
        risk_pregnant=bool(data.get('risk_pregnant')),
        risk_disabled=bool(data.get('risk_disabled')),
        risk_chronic=bool(data.get('risk_chronic')),
        risk_social_vulnerable=bool(data.get('risk_social_vulnerable')),
        risk_lone_elderly=bool(data.get('risk_lone_elderly')),
        risk_needs_care=bool(data.get('risk_needs_care')),
        source=source,
        created_by=user,
        updated_by=user,
    )
    from .primary_care_service import on_population_saved
    sync_meta = on_population_saved(rec, is_new=True)
    rec._primary_care_sync = sync_meta  # noqa: SLF001
    return rec


def upsert_population_from_patient(patient: Patient, user=None) -> PopulationRecord:
    anamnesis_parts = [p for p in [patient.complaints, patient.history] if (p or '').strip()]
    return upsert_population_from_data(
        {
            'registry_number': patient.registry_number,
            'first_name': patient.first_name,
            'last_name': patient.last_name,
            'father_name': patient.father_name,
            'age': patient.age,
            'gender': patient.gender,
            'phone': patient.phone,
            'address': patient.address,
            'region_id': patient.region_id,
            'district_id': patient.district_id,
            'anamnesis': '\n'.join(anamnesis_parts),
        },
        user=user,
        source='patient_auto',
        registry_number=patient.registry_number,
    )


def _parse_bool_cell(raw: Any) -> bool:
    s = str(raw or '').strip().lower()
    return s in ('ha', 'yes', '1', 'true', 'да', '+', 'x')


def _parse_excel_row(row_map: dict[str, Any]) -> dict[str, Any] | None:
    rn_raw = row_map.get('registry_number')
    if not rn_raw:
        return None
    try:
        rn = validate_passport_serial_format(str(rn_raw))
    except Exception:
        rn = normalize_passport_serial(str(rn_raw))
    if not rn:
        return None
    fn = (row_map.get('first_name') or '').strip()
    ln = (row_map.get('last_name') or '').strip()
    if not fn and not ln:
        return None
    brigade_id = None
    brigade_code = (row_map.get('brigade_code') or '').strip()
    if brigade_code:
        from .primary_care_models import MedicalBrigade
        b = MedicalBrigade.objects.filter(code__iexact=brigade_code, is_active=True).first()
        if b:
            brigade_id = b.id
    birth_raw = row_map.get('birth_date')
    birth_date = None
    if birth_raw:
        if hasattr(birth_raw, 'isoformat'):
            birth_date = birth_raw
        else:
            s = str(birth_raw).strip()
            if s:
                from datetime import datetime
                for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
                    try:
                        birth_date = datetime.strptime(s, fmt).date()
                        break
                    except ValueError:
                        continue
    out = {
        'registry_number': rn,
        'first_name': fn or '—',
        'last_name': ln or '—',
        'father_name': (row_map.get('father_name') or '').strip(),
        'age': str(row_map.get('age') or '').strip(),
        'gender': _map_gender(str(row_map.get('gender') or '')),
        'phone': normalize_patient_phone(str(row_map.get('phone') or '')) or str(row_map.get('phone') or '').strip(),
        'address': (row_map.get('address') or '').strip(),
        'anamnesis': (row_map.get('anamnesis') or '').strip(),
        'birth_date': birth_date,
        'health_group': (row_map.get('health_group') or '').strip(),
        'risk_pregnant': _parse_bool_cell(row_map.get('risk_pregnant')),
        'risk_disabled': _parse_bool_cell(row_map.get('risk_disabled')),
        'risk_chronic': _parse_bool_cell(row_map.get('risk_chronic')),
        'risk_social_vulnerable': _parse_bool_cell(row_map.get('risk_social_vulnerable')),
        'risk_lone_elderly': _parse_bool_cell(row_map.get('risk_lone_elderly')),
        'risk_needs_care': _parse_bool_cell(row_map.get('risk_needs_care')),
    }
    if brigade_id:
        out['brigade'] = brigade_id
    return out


def import_population_excel(file_obj, *, user=None) -> dict[str, int]:
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return {'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0}

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        key = _HEADER_ALIASES.get(_normalize_header(str(cell or '')))
        if key and key != 'row_num':
            col_map[idx] = key

    created = updated = skipped = errors = 0
    for row in rows:
        if not row or all(c is None or str(c).strip() == '' for c in row):
            continue
        row_map: dict[str, Any] = {}
        for idx, field in col_map.items():
            if idx < len(row):
                row_map[field] = row[idx]
        parsed = _parse_excel_row(row_map)
        if not parsed:
            skipped += 1
            continue
        try:
            existing = find_population_by_registry(parsed['registry_number'])
            if existing and user and not _population_owned_by_user(user, existing):
                errors += 1
                continue
            upsert_population_from_data(parsed, user=user, source='excel')
            if existing:
                updated += 1
            else:
                created += 1
        except Exception:
            errors += 1
    wb.close()
    return {'created': created, 'updated': updated, 'skipped': skipped, 'errors': errors}


def export_population_excel(*, user=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Aholi'
    header_font = Font(bold=True)
    for col, title in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    qs = PopulationRecord.objects.select_related('brigade').all()
    if user:
        from .primary_care_access import population_for_user
        qs = population_for_user(user)
    for i, rec in enumerate(qs.order_by('last_name', 'first_name'), start=2):
        gender_label = dict(PopulationRecord.GENDER_CHOICES).get(rec.gender, rec.gender)
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=rec.first_name)
        ws.cell(row=i, column=3, value=rec.last_name)
        ws.cell(row=i, column=4, value=rec.father_name)
        ws.cell(row=i, column=5, value=rec.age)
        ws.cell(row=i, column=6, value=gender_label)
        ws.cell(row=i, column=7, value=rec.phone)
        ws.cell(row=i, column=8, value=rec.registry_number)
        ws.cell(row=i, column=9, value=rec.address)
        ws.cell(row=i, column=10, value=rec.anamnesis)
        ws.cell(row=i, column=11, value=rec.birth_date.isoformat() if rec.birth_date else '')
        ws.cell(row=i, column=12, value=rec.health_group or '')
        ws.cell(row=i, column=13, value=rec.brigade.code if rec.brigade else '')
        ws.cell(row=i, column=14, value='ha' if rec.risk_pregnant else '')
        ws.cell(row=i, column=15, value='ha' if rec.risk_disabled else '')
        ws.cell(row=i, column=16, value='ha' if rec.risk_chronic else '')
        ws.cell(row=i, column=17, value='ha' if rec.risk_social_vulnerable else '')
        ws.cell(row=i, column=18, value='ha' if rec.risk_lone_elderly else '')
        ws.cell(row=i, column=19, value='ha' if rec.risk_needs_care else '')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_population_template_excel() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Aholi'
    header_font = Font(bold=True)
    for col, title in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
    ws.cell(row=2, column=2, value='Ali')
    ws.cell(row=2, column=3, value='Valiyev')
    ws.cell(row=2, column=4, value='Vali o\'g\'li')
    ws.cell(row=2, column=5, value='35')
    ws.cell(row=2, column=6, value='Erkak')
    ws.cell(row=2, column=7, value='+998901234567')
    ws.cell(row=2, column=8, value='AB1234567')
    ws.cell(row=2, column=9, value='Andijon viloyati, Andijon tumani')
    ws.cell(row=2, column=10, value='')
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
