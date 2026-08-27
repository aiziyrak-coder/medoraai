"""So'x KTMP Excel (Bemorlar) — qatorlarni Patient/Population/ko'rik formatiga o'girish."""
from __future__ import annotations

import re
from collections import Counter
from datetime import date
from typing import Any

from openpyxl import load_workbook

SOX_REGION_ID = '12'  # Farg'ona viloyati
SOX_DISTRICT_ID = '180'  # So'x tumani
DEFAULT_CHECKUP_DATE = date(2026, 7, 29)

_APOS = ("'", "'", '`', '´', 'ʼ', 'ʻ', '’', '‘', '‛', '′', '＇', '\u02bb', '\u02bc')


def _norm_apostrophe(s: str) -> str:
    out = (s or '').upper()
    for ch in _APOS:
        out = out.replace(ch, "'")
    return out


def normalize_medcard(value: Any) -> str:
    return re.sub(r'[\s\-]', '', str(value or '').strip().upper())


def parse_health_group(raw: Any) -> str:
    m = re.search(r'([1-4])', str(raw or ''))
    return m.group(1) if m else ''


def parse_age(raw: Any) -> str:
    if raw is None:
        return ''
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return str(int(raw))
    m = re.search(r'\d+', str(raw).strip())
    return m.group() if m else ''


def parse_full_name(raw: Any) -> tuple[str, str, str]:
    """Familiya takrorlangan 'LAST LAST FIRST …' formatini ajratadi."""
    toks = [t for t in str(raw or '').split() if t]
    if not toks:
        return '—', '—', ''
    if len(toks) >= 2 and toks[0].upper() == toks[1].upper():
        last = toks[0]
        rest = toks[2:]
    else:
        last = toks[0]
        rest = toks[1:]
    first = rest[0] if rest else '—'
    father = ' '.join(rest[1:]) if len(rest) > 1 else ''
    return last, first, father


def infer_gender(first_name: str, father_name: str, last_name: str) -> str:
    blob = _norm_apostrophe(f'{first_name} {father_name} {last_name}')
    if re.search(r"QIZI\b", blob) or re.search(r"(OVA|EVA|YEVA|OVNA|EVNA)\b", blob):
        return 'female'
    if re.search(r"O'?G'?LI\b", blob) or re.search(r"(OVICH|EVICH|OGLI)\b", blob):
        return 'male'
    last_u = _norm_apostrophe(last_name).rstrip('.')
    if last_u.endswith(('OVA', 'EVA', 'YEVA')):
        return 'female'
    if last_u.endswith(('OV', 'EV', 'YEV')):
        return 'male'
    first_u = _norm_apostrophe(first_name)
    if first_u.endswith(('XON', 'KHON', 'BEGIM', 'OY')):
        return 'female'
    return ''


def parse_disability(raw: Any) -> tuple[bool, str]:
    text = str(raw or '').strip()
    if not text:
        return False, ''
    low = text.lower()
    if low in ('нет', 'нет.', 'yoq', "yo'q", 'yoʻq', 'no', '-', '0', 'йўқ'):
        return False, ''
    return True, text


def parse_disability_group(raw: Any) -> tuple[bool, str]:
    """Nogironlik matnini guruhga (I/II/III/childhood) o'girish."""
    text = str(raw or '').strip()
    if not text:
        return False, ''
    low = text.lower()
    if low in ('нет', 'нет.', 'yoq', "yo'q", 'yoʻq', 'no', '-', '0', 'йўқ'):
        return False, ''
    if re.search(r'дет', low) or 'bolalik' in low:
        return True, 'childhood'
    if re.search(r'\biii\b', low) or '3 группа' in low:
        return True, 'III'
    if re.search(r'\bii\b', low) or '2 группа' in low:
        return True, 'II'
    if re.search(r'\bi\b', low) or '1 группа' in low:
        return True, 'I'
    return True, text[:20]


def parse_icd(raw: Any) -> str:
    return str(raw or '').strip()[:20]


def build_clinical_notes(*, health_group: str, icd: str, disability: str, medcard: str) -> str:
    parts = [
        f"Manba: So'x KTMP Excel (ko'rik {DEFAULT_CHECKUP_DATE.isoformat()})",
        f"Medkarta: {medcard}",
    ]
    if health_group:
        parts.append(f"Sog'liq guruhi: {health_group}")
    if icd:
        parts.append(f"D-hisob: {icd}")
    if disability:
        parts.append(f"Nogironlik: {disability}")
    return '\n'.join(parts)


def iter_sox_rows(path: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Excel qatorlari. Bir medkarta uchun oxirgi qator qoladi."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    by_card: dict[str, dict[str, Any]] = {}
    skipped_empty = 0
    duplicate_extra = 0
    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        card = normalize_medcard(row[1] if len(row) > 1 else '')
        name = str(row[2] if len(row) > 2 else '').strip()
        if not card or not name:
            skipped_empty += 1
            continue
        if card in by_card:
            duplicate_extra += 1
        last_name, first_name, father_name = parse_full_name(name)
        age = parse_age(row[3] if len(row) > 3 else None)
        health_group = parse_health_group(row[4] if len(row) > 4 else None)
        icd = parse_icd(row[5] if len(row) > 5 else None)
        disabled, disability_text = parse_disability(row[6] if len(row) > 6 else None)
        _, disability_group = parse_disability_group(disability_text if disabled else '')
        gender = infer_gender(first_name, father_name, last_name)
        by_card[card] = {
            'registry_number': card,
            'medical_card_number': card,
            'last_name': last_name,
            'first_name': first_name,
            'father_name': father_name,
            'age': age,
            'gender': gender,
            'health_group': health_group,
            'icd10_code': icd,
            'dispensary_icd_code': icd,
            'disability_group': disability_group,
            'risk_disabled': disabled,
            'disability_text': disability_text,
            'risk_chronic': bool(icd),
            'dispensary_registered': bool(icd),
            'region_id': SOX_REGION_ID,
            'district_id': SOX_DISTRICT_ID,
            'notes': build_clinical_notes(
                health_group=health_group,
                icd=icd,
                disability=disability_text,
                medcard=card,
            ),
        }
    wb.close()
    records = list(by_card.values())
    return records, {
        'excel_rows': total_rows,
        'unique_cards': len(records),
        'duplicate_extra': duplicate_extra,
        'skipped_empty': skipped_empty,
    }


def summarize_records(records: list[dict[str, Any]]) -> dict[str, Any]:
    hg = Counter(r['health_group'] or '' for r in records)
    gender = Counter(r['gender'] or '' for r in records)
    return {
        'patients': len(records),
        'health_groups': dict(hg),
        'with_health_group': sum(1 for r in records if r['health_group']),
        'dispensary': sum(1 for r in records if r['icd10_code']),
        'disabled': sum(1 for r in records if r['risk_disabled']),
        'missing_age': sum(1 for r in records if not r['age']),
        'gender': dict(gender),
        'top_icd': Counter(r['icd10_code'] for r in records if r['icd10_code']).most_common(12),
    }
