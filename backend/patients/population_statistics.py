"""Bemorlar statistikasi — filtrlar va kesimlar (barchasi SQL darajasida)."""
from __future__ import annotations

import io
import logging
import re
from collections import Counter
from typing import Any

from django.db.models import CharField, Count, Exists, OuterRef, Q, QuerySet, Subquery, Value
from django.db.models.functions import Coalesce, NullIf, Upper
from openpyxl import Workbook
from openpyxl.styles import Font

from .icd10_catalog import ICD10_CHAPTERS, chapter_by_code, chapter_for_icd, normalize_icd_code
from .models import PopulationRecord
from .primary_care_models import DispensaryRecord

logger = logging.getLogger(__name__)

AGE_BUCKETS = (
    ('0-1', 0, 1),
    ('2-17', 2, 17),
    ('18-29', 18, 29),
    ('30-44', 30, 44),
    ('45-59', 45, 59),
    ('60-74', 60, 74),
    ('75+', 75, 200),
)

# SQLite da SQLITE_MAX_VARIABLE_NUMBER = 999 — IN ro'yxati shundan oshmasligi kerak
_MAX_IN_PARAMS = 900
_MAX_DISTINCT_AGE_VALUES = 2000
# Excel eksport uchun qattiq chegara (xotira va gunicorn timeout)
EXPORT_MAX_ROWS = 50000
# Statistika keshi (sekund)
STATISTICS_CACHE_TTL = 120


def parse_age_int(age_raw: str) -> int | None:
    m = re.search(r'\d+', str(age_raw or ''))
    if not m:
        return None
    try:
        return int(m.group())
    except ValueError:
        return None


def age_bucket(age_raw: str) -> str:
    age = parse_age_int(age_raw)
    if age is None:
        return 'unknown'
    for label, lo, hi in AGE_BUCKETS:
        if lo <= age <= hi:
            return label
    return 'unknown'


def _effective_disability_group(group: str, risk_disabled: bool) -> str:
    g = (group or '').strip()
    if g:
        return g
    if risk_disabled:
        return 'unknown'
    return ''


def annotate_effective_icd(qs: QuerySet) -> QuerySet:
    """
    D-hisob kodi: PopulationRecord dagi kod, bo'lmasa oxirgi faol dispanser yozuvidan.
    Subquery orqali — ilgari bu yerda barcha id lar ro'yxati bilan IN so'rovi bor edi
    (SQLite da 999 ta o'zgaruvchi chegarasi va katta xotira).
    """
    if 'effective_icd' in qs.query.annotations:
        return qs
    latest_disp = (
        DispensaryRecord.objects.filter(population_id=OuterRef('pk'), is_active=True)
        .exclude(icd10_code='')
        .order_by('-registered_date')
        .values('icd10_code')[:1]
    )
    return qs.annotate(
        effective_icd=Upper(
            Coalesce(
                NullIf('dispensary_icd_code', Value('')),
                Subquery(latest_disp),
                Value(''),
                output_field=CharField(),
            )
        ),
    )


def _chapter_range(chapter) -> tuple[str, str]:
    """XKT-10 sinfi uchun kod oralig'i: [start, end_exclusive) — matnli taqqoslash uchun."""
    start = f'{chapter.letter_start}{chapter.num_start:02d}'
    if chapter.num_end >= 99:
        end_exclusive = f'{chr(ord(chapter.letter_end) + 1)}00'
    else:
        end_exclusive = f'{chapter.letter_end}{chapter.num_end + 1:02d}'
    return start, end_exclusive


def _apply_age_value_filter(qs: QuerySet, matches, fallback: list[str]) -> QuerySet:
    """
    Yosh — matn maydoni ("45", "45 yosh"). Mavjud turli qiymatlar ro'yxati kichik
    (bitta yengil DISTINCT so'rov), shuning uchun mos qiymatlarni Python da tanlab,
    filtrni SQL ga IN ko'rinishida beramiz — jadval Python ga yuklanmaydi.
    """
    distinct = list(
        qs.order_by().values_list('age', flat=True).distinct()[:_MAX_DISTINCT_AGE_VALUES]
    )
    match = [v for v in distinct if matches(v)]
    if len(distinct) < _MAX_DISTINCT_AGE_VALUES:
        if len(match) <= _MAX_IN_PARAMS:
            return qs.filter(age__in=match)
        match_set = set(match)
        skip = [v for v in distinct if v not in match_set]
        if len(skip) <= _MAX_IN_PARAMS:
            return qs.exclude(age__in=skip)
    logger.warning(
        "Yosh filtri: juda ko'p turli qiymat (%s) — faqat sof raqamli yoshlar bo'yicha filtrlanadi",
        len(distinct),
    )
    return qs.filter(age__in=fallback)


def _apply_age_range_filter(qs: QuerySet, lo: int, hi: int) -> QuerySet:
    def _matches(value: str) -> bool:
        a = parse_age_int(value)
        return a is not None and lo <= a <= hi

    return _apply_age_value_filter(
        qs, _matches, [str(a) for a in range(max(lo, 0), min(hi, 200) + 1)],
    )


def _apply_age_bucket_filter(qs: QuerySet, label: str) -> QuerySet:
    fallback = [
        str(a) for a in range(0, 201)
        if age_bucket(str(a)) == label
    ]
    return _apply_age_value_filter(qs, lambda v: age_bucket(v) == label, fallback)


def filter_population_queryset(qs: QuerySet, params: dict[str, Any]) -> QuerySet:
    """Filtrlarni to'liq SQL darajasida qo'llaydi (natija — queryset, ro'yxat emas)."""
    region_id = (params.get('region_id') or '').strip()
    district_id = (params.get('district_id') or '').strip()
    health_group = (params.get('health_group') or '').strip()
    gender = (params.get('gender') or '').strip()
    q = (params.get('q') or params.get('search') or '').strip()
    disease_chapter = (params.get('disease_chapter') or params.get('disease_type') or '').strip()
    icd_code = normalize_icd_code(params.get('icd_code') or params.get('dispensary_icd') or '')
    age_group = (params.get('age_group') or '').strip()
    age_min = params.get('age_min')
    age_max = params.get('age_max')
    disability = (params.get('disability') or '').strip().lower()
    disability_group = (params.get('disability_group') or '').strip()
    dispensary = (params.get('dispensary') or params.get('d_account') or '').strip().lower()

    qs = annotate_effective_icd(qs)

    brigade_id = params.get('brigade_id')
    if brigade_id:
        qs = qs.filter(brigade_id=int(brigade_id))
    if region_id:
        qs = qs.filter(region_id=region_id)
    if district_id:
        qs = qs.filter(district_id=district_id)
    if health_group:
        qs = qs.filter(health_group=health_group)
    if gender in ('male', 'female', 'other'):
        qs = qs.filter(gender=gender)
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(father_name__icontains=q)
            | Q(registry_number__icontains=q)
            | Q(medical_card_number__icontains=q)
            | Q(phone__icontains=q)
        )

    if disability in ('yes', 'true', '1', 'ha', 'bor'):
        qs = qs.filter(Q(risk_disabled=True) | ~Q(disability_group=''))
    elif disability in ('no', 'false', '0', 'yoq', "yo'q", 'yoq'):
        qs = qs.filter(risk_disabled=False, disability_group='')

    if disability_group:
        qs = qs.filter(disability_group__iexact=disability_group)

    # Dispanser bog'lanishi Exists() bilan — JOIN dublikatlari va .distinct() kerak emas
    active_disp = DispensaryRecord.objects.filter(population_id=OuterRef('pk'), is_active=True)
    if dispensary in ('yes', 'true', '1', 'ha', 'bor'):
        qs = qs.filter(
            Q(dispensary_registered=True)
            | ~Q(dispensary_icd_code='')
            | Exists(active_disp),
        )
    elif dispensary in ('no', 'false', '0', 'yoq', "yo'q", 'yoq'):
        qs = qs.filter(dispensary_registered=False, dispensary_icd_code='').filter(
            ~Exists(active_disp),
        )

    if icd_code:
        qs = qs.filter(
            Q(dispensary_icd_code__iexact=icd_code)
            | Q(dispensary_icd_code__istartswith=icd_code)
            | Exists(active_disp.filter(icd10_code__iexact=icd_code)),
        )

    if age_group:
        qs = _apply_age_bucket_filter(qs, age_group)
    elif age_min not in (None, '') or age_max not in (None, ''):
        lo = int(age_min) if age_min not in (None, '') else 0
        hi = int(age_max) if age_max not in (None, '') else 200
        qs = _apply_age_range_filter(qs, lo, hi)

    if disease_chapter:
        ch = chapter_by_code(disease_chapter)
        if ch:
            start, end_exclusive = _chapter_range(ch)
            qs = qs.filter(effective_icd__gte=start, effective_icd__lt=end_exclusive)

    return qs


def compute_population_statistics(qs: QuerySet, *, lang: str = 'uz') -> dict[str, Any]:
    """Kesimlarni DB agregatlari bilan hisoblaydi — qatorlar Python ga yuklanmaydi."""
    qs = annotate_effective_icd(qs).order_by()

    by_district: Counter[str] = Counter()
    by_region: Counter[str] = Counter()
    by_age: Counter[str] = Counter()
    by_health: Counter[str] = Counter()
    by_gender: Counter[str] = Counter()
    by_disability_group: Counter[str] = Counter()
    by_disease_chapter: Counter[str] = Counter()
    by_icd: Counter[str] = Counter()
    disabled_count = 0
    dispensary_count = 0

    region_names, district_meta = _region_names()

    total = qs.count()

    for row in qs.values('region_id', 'district_id').annotate(c=Count('id', distinct=True)):
        rid = str(row['region_id'] or '')
        did = str(row['district_id'] or '')
        by_region[region_names.get(rid, rid or '—')] += row['c']
        by_district[district_meta.get(did, {}).get('district_name', did or '—')] += row['c']

    for row in qs.values('age').annotate(c=Count('id', distinct=True)):
        by_age[age_bucket(row['age'])] += row['c']

    for row in qs.values('health_group').annotate(c=Count('id', distinct=True)):
        by_health[row['health_group'] or '—'] += row['c']

    for row in qs.values('gender').annotate(c=Count('id', distinct=True)):
        by_gender[row['gender'] or '—'] += row['c']

    for row in qs.values('disability_group', 'risk_disabled').annotate(c=Count('id', distinct=True)):
        dg = _effective_disability_group(row['disability_group'], row['risk_disabled'])
        if not dg:
            continue
        disabled_count += row['c']
        by_disability_group[dg] += row['c']

    for row in qs.values('effective_icd', 'dispensary_registered').annotate(c=Count('id', distinct=True)):
        icd = normalize_icd_code(row['effective_icd'] or '')
        if icd or row['dispensary_registered']:
            dispensary_count += row['c']
        if icd:
            by_icd[icd] += row['c']
            ch = chapter_for_icd(icd)
            if ch:
                label = ch.name_ru if lang == 'ru' else ch.name_uz
                by_disease_chapter[label] += row['c']

    name_key = 'name_ru' if lang == 'ru' else 'name_uz'
    chapters_meta = [
        {'code': c.code, 'range': c.range_label, 'label': getattr(c, name_key)}
        for c in ICD10_CHAPTERS
    ]

    return {
        'total': total,
        'disabled_total': disabled_count,
        'dispensary_total': dispensary_count,
        'by_region': _counter_list(by_region),
        'by_district': _counter_list(by_district),
        'by_age_group': _counter_list(by_age, order=[b[0] for b in AGE_BUCKETS] + ['unknown']),
        'by_health_group': _counter_list(by_health),
        'by_gender': _counter_list(by_gender),
        'by_disability_group': _counter_list(by_disability_group),
        'by_disease_chapter': _counter_list(by_disease_chapter),
        'top_icd_codes': [{'code': k, 'count': v} for k, v in by_icd.most_common(25)],
        'icd_chapters': chapters_meta,
        'age_buckets': [b[0] for b in AGE_BUCKETS],
    }


def _counter_list(c: Counter, order: list[str] | None = None) -> list[dict[str, Any]]:
    if order:
        keys = [k for k in order if k in c] + [k for k in c if k not in order]
    else:
        keys = [k for k, _ in c.most_common()]
    return [{'label': k, 'count': c[k]} for k in keys]


def _region_names() -> tuple[dict[str, str], dict[str, dict]]:
    from .address_data import load_address_catalog
    catalog = load_address_catalog()
    region_names = {r['id']: r['name_uz'] for r in catalog['regions']}
    district_meta: dict[str, dict] = {}
    for r in catalog['regions']:
        for d in r.get('districts', []):
            district_meta[str(d['id'])] = {
                'district_name': d['name_uz'],
                'region_id': r['id'],
                'region_name': r['name_uz'],
            }
    return region_names, district_meta


def export_statistics_excel(qs: QuerySet) -> bytes:
    """
    Excel eksport — saralash SQL da, qatorlar iterator bilan oqim shaklida o'qiladi
    (butun jadval Python ro'yxatiga yuklanmaydi) va EXPORT_MAX_ROWS bilan cheklanadi.
    """
    qs = annotate_effective_icd(qs).order_by('last_name', 'first_name')
    region_names, district_meta = _region_names()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bemorlar'
    headers = [
        'Medkarta', 'Familiya', 'Ism', 'Otasining ismi', 'Yosh', 'Jins',
        'Viloyat', 'Tuman', 'Sog\'liq guruhi', 'NKB (D-hisob)', 'Nogironlik guruhi',
    ]
    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold

    rows = qs.values(
        'medical_card_number', 'registry_number', 'last_name', 'first_name', 'father_name',
        'age', 'gender', 'region_id', 'district_id', 'health_group',
        'effective_icd', 'disability_group', 'risk_disabled',
    )[:EXPORT_MAX_ROWS]
    written = 0
    for i, rec in enumerate(rows.iterator(chunk_size=1000), start=2):
        meta = district_meta.get(str(rec['district_id'] or ''), {})
        ws.cell(row=i, column=1, value=rec['medical_card_number'] or rec['registry_number'])
        ws.cell(row=i, column=2, value=rec['last_name'])
        ws.cell(row=i, column=3, value=rec['first_name'])
        ws.cell(row=i, column=4, value=rec['father_name'])
        ws.cell(row=i, column=5, value=rec['age'])
        ws.cell(row=i, column=6, value=rec['gender'])
        ws.cell(row=i, column=7, value=region_names.get(str(rec['region_id'] or ''), ''))
        ws.cell(row=i, column=8, value=meta.get('district_name', ''))
        ws.cell(row=i, column=9, value=rec['health_group'])
        ws.cell(row=i, column=10, value=normalize_icd_code(rec['effective_icd'] or ''))
        ws.cell(row=i, column=11, value=_effective_disability_group(
            rec['disability_group'], rec['risk_disabled'],
        ))
        written += 1
    if written >= EXPORT_MAX_ROWS:
        logger.warning('Statistika eksporti %s qator bilan cheklandi', EXPORT_MAX_ROWS)
        ws.cell(row=written + 2, column=1, value=f'... {EXPORT_MAX_ROWS} qator bilan cheklandi')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def backfill_population_statistics_fields(batch_size: int = 500) -> dict[str, int]:
    updated = 0
    from .sox_excel_import import parse_disability_group

    for pop in PopulationRecord.objects.all().order_by('id').iterator(chunk_size=batch_size):
        changed = False
        if not pop.medical_card_number and pop.registry_number:
            pop.medical_card_number = pop.registry_number
            changed = True
        if not pop.dispensary_icd_code:
            disp = (
                DispensaryRecord.objects.filter(population=pop, is_active=True)
                .order_by('-registered_date')
                .first()
            )
            if disp and disp.icd10_code:
                pop.dispensary_icd_code = normalize_icd_code(disp.icd10_code)
                pop.dispensary_diagnosis = (disp.diagnosis or '')[:255]
                pop.dispensary_registered = True
                changed = True
        if not pop.disability_group and pop.risk_disabled and 'Nogironlik:' in (pop.anamnesis or ''):
            part = pop.anamnesis.split('Nogironlik:', 1)[1].split('\n', 1)[0].strip()
            _, grp = parse_disability_group(part)
            if grp:
                pop.disability_group = grp
                changed = True
        if changed:
            pop.save(update_fields=[
                'medical_card_number', 'dispensary_icd_code', 'dispensary_diagnosis',
                'dispensary_registered', 'disability_group', 'updated_at',
            ])
            updated += 1
    return {'updated': updated}
