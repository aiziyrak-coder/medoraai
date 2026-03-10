"""
Monitoring API – devices, vitals, alarms.
JWT auth; monitoring role required for write; clinic/monitoring for read.
Ingest endpoint uses API key for gateway (no JWT).
"""
import logging
from django.conf import settings
from django.utils import timezone
from rest_framework import status, permissions, generics
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView

from django.contrib.auth import get_user_model

from .models import (
    Ward, Room, Device, PatientMonitor, VitalReading, AlarmThreshold, Alarm,
    MonitoringAuditLog, MonitoringNote,
    MonitoringMedication, MonitoringLabResult, FamilyViewToken,
)
from .services import evaluate_vital_alarms, get_vital_status, effective_device_status
from .serializers import (
    WardSerializer, RoomSerializer, DeviceSerializer, DeviceRegisterSerializer,
    PatientMonitorSerializer, VitalReadingSerializer, AlarmThresholdSerializer, AlarmSerializer,
    MonitoringAuditLogSerializer, MonitoringNoteSerializer,
)

logger = logging.getLogger(__name__)

INGEST_API_KEY = getattr(settings, 'MONITORING_INGEST_API_KEY', '')


def is_monitoring_user(request):
    """Faqat monitoring yoki clinic rolida boʻlsa True."""
    user = getattr(request, 'user', None)
    if not user or not user.is_authenticated:
        return False
    return getattr(user, 'role', None) in ('monitoring', 'clinic', 'doctor', 'staff')


class MonitoringPermission(permissions.BasePermission):
    """Monitoring API: faqat autentifikatsiya qilingan va monitoring/clinic/doctor/staff."""
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated and is_monitoring_user(request)


class WardListCreate(generics.ListCreateAPIView):
    permission_classes = [MonitoringPermission]
    serializer_class = WardSerializer
    queryset = Ward.objects.filter(is_active=True)


class WardDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [MonitoringPermission]
    serializer_class = WardSerializer
    queryset = Ward.objects.filter(is_active=True)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()


class RoomListCreate(generics.ListCreateAPIView):
    permission_classes = [MonitoringPermission]
    serializer_class = RoomSerializer
    queryset = Room.objects.filter(is_active=True).select_related('ward')


class RoomDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [MonitoringPermission]
    serializer_class = RoomSerializer
    queryset = Room.objects.filter(is_active=True).select_related('ward')

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()


class DeviceList(generics.ListAPIView):
    permission_classes = [MonitoringPermission]
    serializer_class = DeviceSerializer
    queryset = Device.objects.filter(is_active=True).select_related('room')

    def get_queryset(self):
        qs = super().get_queryset()
        room_id = self.request.query_params.get('room')
        if room_id:
            qs = qs.filter(room_id=room_id)
        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        return qs


class DeviceRegister(APIView):
    """POST /api/v1/devices/register – yangi monitorni tizimga qoʻshish."""
    permission_classes = [MonitoringPermission]

    def post(self, request):
        ser = DeviceRegisterSerializer(data=request.data)
        if not ser.is_valid():
            return Response(ser.errors, status=status.HTTP_400_BAD_REQUEST)
        device = ser.save()
        return Response(DeviceSerializer(device).data, status=status.HTTP_201_CREATED)


class DeviceStatus(APIView):
    """GET /api/v1/devices/status – barcha monitorlarning onlayn/offlayn holati."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        devices = Device.objects.filter(is_active=True).select_related('room')
        room_id = request.query_params.get('room')
        if room_id:
            devices = devices.filter(room_id=room_id)
        data = DeviceSerializer(devices, many=True).data
        return Response({'success': True, 'data': data})


class DeviceDetail(generics.RetrieveUpdateDestroyAPIView):
    """GET/PATCH/DELETE device – tahrirlash yoki o'chirish (soft: is_active=False)."""
    permission_classes = [MonitoringPermission]
    serializer_class = DeviceSerializer
    queryset = Device.objects.filter(is_active=True).select_related('room')

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()


class PatientMonitorListCreate(generics.ListCreateAPIView):
    permission_classes = [MonitoringPermission]
    serializer_class = PatientMonitorSerializer
    queryset = PatientMonitor.objects.filter(is_active=True).select_related('device', 'room', 'assigned_to')

    def get_queryset(self):
        qs = super().get_queryset()
        room_id = self.request.query_params.get('room')
        if room_id:
            qs = qs.filter(room_id=room_id)
        return qs


class PatientMonitorDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [MonitoringPermission]
    serializer_class = PatientMonitorSerializer
    queryset = PatientMonitor.objects.filter(is_active=True).select_related('device', 'room', 'assigned_to')


class VitalReadingList(generics.ListAPIView):
    """GET vitals – patient_monitor yoki device boʻyicha, vaqt oraligʻi."""
    permission_classes = [MonitoringPermission]
    serializer_class = VitalReadingSerializer

    def get_queryset(self):
        qs = VitalReading.objects.all().select_related('patient_monitor', 'patient_monitor__device', 'patient_monitor__room')
        pm_id = self.request.query_params.get('patient_monitor_id')
        if pm_id:
            qs = qs.filter(patient_monitor_id=pm_id)
        room_id = self.request.query_params.get('room_id')
        if room_id:
            qs = qs.filter(patient_monitor__room_id=room_id)
        from_ts = self.request.query_params.get('from')
        to_ts = self.request.query_params.get('to')
        if from_ts:
            qs = qs.filter(timestamp__gte=from_ts)
        if to_ts:
            qs = qs.filter(timestamp__lte=to_ts)
        limit = min(int(self.request.query_params.get('limit', 100)), 500)
        return qs.order_by('-timestamp')[:limit]


class VitalReadingCreate(APIView):
    """POST vitals – gateway dan kelgan maʼlumotlarni yozish (yoki mock)."""
    permission_classes = [MonitoringPermission]

    def post(self, request):
        patient_monitor_id = request.data.get('patient_monitor_id')
        if not patient_monitor_id:
            return Response({'success': False, 'error': 'patient_monitor_id required'}, status=400)
        try:
            pm = PatientMonitor.objects.get(pk=patient_monitor_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'PatientMonitor not found'}, status=404)
        ts = request.data.get('timestamp') or timezone.now()
        reading = VitalReading.objects.create(
            patient_monitor=pm,
            timestamp=ts,
            heart_rate=request.data.get('heart_rate'),
            spo2=request.data.get('spo2'),
            nibp_systolic=request.data.get('nibp_systolic'),
            nibp_diastolic=request.data.get('nibp_diastolic'),
            respiration_rate=request.data.get('respiration_rate'),
            temperature=request.data.get('temperature'),
            raw_payload=request.data.get('raw_payload', {}),
        )
        # Update device last_seen
        pm.device.status = 'online'
        pm.device.last_seen_at = timezone.now()
        pm.device.save(update_fields=['status', 'last_seen_at'])
        # TODO: Alarm engine – threshold tekshirish va Alarm yaratish
        return Response(VitalReadingSerializer(reading).data, status=status.HTTP_201_CREATED)


class IngestPermission(permissions.BasePermission):
    """Gateway ingest: X-API-Key header must match MONITORING_INGEST_API_KEY."""

    def has_permission(self, request, view):
        if not INGEST_API_KEY:
            return False
        key = request.headers.get('X-API-Key') or request.headers.get('Authorization', '').replace('Bearer ', '')
        return key == INGEST_API_KEY


class GatewayMonitors(APIView):
    """
    GET /api/monitoring/gateway-monitors/ – gateway ulanish ro'yxati (X-API-Key).
    Faqat host va port to'ldirilgan qurilmalar (TCP ulanish uchun).
    """
    permission_classes = [IngestPermission]
    authentication_classes = []

    def get(self, request):
        devices = Device.objects.filter(is_active=True).exclude(host='').exclude(port__isnull=True)
        monitors = [
            {'device_id': d.serial_number, 'host': (d.host or '').strip(), 'port': int(d.port)}
            for d in devices
            if (d.host or '').strip() and d.port
        ]
        return Response({'success': True, 'monitors': monitors})


class IngestVitals(APIView):
    """
    POST /api/monitoring/ingest/ – gateway dan real-time vitals (API key).
    Body: device_id (Device.serial_number), heart_rate, spo2, bp_sys, bp_dia, respiration, temperature, timestamp.
    """
    permission_classes = [IngestPermission]
    authentication_classes = []  # No JWT

    def post(self, request):
        device_id = (request.data.get('device_id') or '').strip()
        if not device_id:
            return Response({'success': False, 'error': 'device_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            device = Device.objects.select_related('patient_monitor').get(serial_number=device_id, is_active=True)
        except Device.DoesNotExist:
            return Response({'success': False, 'error': f'Device {device_id} not found'}, status=status.HTTP_404_NOT_FOUND)
        try:
            pm = device.patient_monitor
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': f'No patient monitor for device {device_id}'}, status=status.HTTP_404_NOT_FOUND)
        if not pm.is_active:
            return Response({'success': False, 'error': 'Patient monitor inactive'}, status=status.HTTP_400_BAD_REQUEST)

        ts = request.data.get('timestamp') or timezone.now()
        reading = VitalReading.objects.create(
            patient_monitor=pm,
            timestamp=ts,
            heart_rate=request.data.get('heart_rate'),
            spo2=request.data.get('spo2'),
            nibp_systolic=request.data.get('nibp_systolic') or request.data.get('bp_sys'),
            nibp_diastolic=request.data.get('nibp_diastolic') or request.data.get('bp_dia'),
            respiration_rate=request.data.get('respiration_rate') or request.data.get('respiration'),
            temperature=request.data.get('temperature'),
            raw_payload=request.data.get('raw_payload', {}),
        )
        device.status = 'online'
        device.last_seen_at = timezone.now()
        device.save(update_fields=['status', 'last_seen_at'])
        # Evaluate clinical alarms (SpO2, HR, BP) and create Alarm records
        evaluate_vital_alarms(reading)
        return Response({
            'success': True,
            'reading_id': reading.id,
            'patient_monitor_id': pm.id,
        }, status=status.HTTP_201_CREATED)


class AlarmList(generics.ListAPIView):
    permission_classes = [MonitoringPermission]
    serializer_class = AlarmSerializer
    queryset = Alarm.objects.all().select_related('patient_monitor', 'patient_monitor__device', 'patient_monitor__room')

    def get_queryset(self):
        qs = super().get_queryset()
        acknowledged = self.request.query_params.get('acknowledged')
        if acknowledged is not None:
            if acknowledged.lower() == 'true':
                qs = qs.exclude(acknowledged_at__isnull=True)
            else:
                qs = qs.filter(acknowledged_at__isnull=True)
        patient_monitor_id = self.request.query_params.get('patient_monitor_id')
        if patient_monitor_id:
            qs = qs.filter(patient_monitor_id=patient_monitor_id)
        return qs.order_by('-created_at')[:100]


class AlarmAcknowledge(APIView):
    permission_classes = [MonitoringPermission]

    def post(self, request, pk):
        try:
            alarm = Alarm.objects.get(pk=pk)
        except Alarm.DoesNotExist:
            return Response({'success': False, 'error': 'Alarm not found'}, status=404)
        if alarm.acknowledged_at:
            return Response(AlarmSerializer(alarm).data)
        alarm.acknowledged_at = timezone.now()
        alarm.acknowledged_by = request.user
        alarm.save(update_fields=['acknowledged_at', 'acknowledged_by'])
        MonitoringAuditLog.objects.create(
            patient_monitor=alarm.patient_monitor,
            action='alarm_ack',
            user=request.user,
            details={'alarm_id': alarm.pk, 'param': alarm.param, 'value': alarm.value},
        )
        return Response(AlarmSerializer(alarm).data)


class AlarmThresholdListCreate(generics.ListCreateAPIView):
    permission_classes = [MonitoringPermission]
    serializer_class = AlarmThresholdSerializer
    queryset = AlarmThreshold.objects.filter(is_active=True)

    def get_queryset(self):
        qs = super().get_queryset()
        pm_id = self.request.query_params.get('patient_monitor_id')
        if pm_id:
            qs = qs.filter(patient_monitor_id=pm_id)
        return qs


def _vital_status_from_reading(last):
    """Return 'normal'|'warning'|'critical' from last VitalReading."""
    if not last:
        return 'normal'
    return get_vital_status({
        'heart_rate': last.heart_rate,
        'spo2': last.spo2,
        'nibp_systolic': last.nibp_systolic,
        'nibp_diastolic': last.nibp_diastolic,
    })


def _ews_from_reading(last):
    """
    Early Warning Score (MEWS-style) from last VitalReading.
    Returns (score: int 0-15+, level: 'past'|'o'rta'|'yuqori').
    """
    if not last:
        return 0, 'past'
    score = 0
    # HR
    hr = last.heart_rate
    if hr is not None:
        if hr < 40 or hr >= 130:
            score += 3
        elif hr < 50 or (hr >= 100 and hr < 130):
            score += 2
        elif hr >= 40 and hr < 50 or (hr >= 110 and hr < 130):
            score += 1
    # SpO2
    spo2 = last.spo2
    if spo2 is not None:
        if spo2 < 85:
            score += 3
        elif spo2 < 90:
            score += 2
        elif spo2 < 95:
            score += 1
    # Systolic BP
    sys_ = last.nibp_systolic
    if sys_ is not None:
        if sys_ < 70:
            score += 3
        elif sys_ < 80:
            score += 2
        elif sys_ < 90 or sys_ >= 180:
            score += 1
    # Respiration
    rr = last.respiration_rate
    if rr is not None:
        if rr < 9 or rr >= 25:
            score += 2
        elif rr >= 21 or (rr >= 15 and rr < 21):
            score += 1
    # Temperature
    try:
        temp = float(last.temperature) if last.temperature is not None else None
    except (TypeError, ValueError):
        temp = None
    if temp is not None:
        if temp < 35:
            score += 2
        elif temp >= 38.5:
            score += 1
    if score >= 7:
        level = 'yuqori'
    elif score >= 4:
        level = "o'rta"
    else:
        level = 'past'
    return score, level


def _fall_risk_from_pm(pm, last):
    """Oddiy yiqilish xavfi: yosh, vitals (gipotensiya). past/o'rta/yuqori."""
    risk = 0
    if pm.age is not None:
        if pm.age >= 80:
            risk += 2
        elif pm.age >= 65:
            risk += 1
    if last and last.nibp_systolic is not None and last.nibp_systolic < 90:
        risk += 2
    if last and last.heart_rate is not None and (last.heart_rate > 120 or last.heart_rate < 50):
        risk += 1
    if risk >= 4:
        return 'yuqori'
    if risk >= 2:
        return "o'rta"
    return 'past'


def _pressure_risk_from_pm(pm, last):
    """Soddalashtirilgan bosim yarasi xavfi (harakatlanish yo'q – faqat vitals). past/o'rta/yuqori."""
    risk = 0
    if pm.age is not None and pm.age >= 70:
        risk += 1
    if last and last.spo2 is not None and last.spo2 < 92:
        risk += 1
    if last and last.nibp_systolic is not None and last.nibp_systolic < 100:
        risk += 1
    if risk >= 2:
        return 'yuqori'
    if risk >= 1:
        return "o'rta"
    return 'past'


class DashboardSummary(APIView):
    """GET /api/monitoring/dashboard/ – grid, summary, filter (room_id, status, search)."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        monitors = PatientMonitor.objects.filter(is_active=True).select_related('device', 'room', 'room__ward', 'assigned_to')
        room_id = request.query_params.get('room_id')
        if room_id:
            monitors = monitors.filter(room_id=room_id)
        ward_id = request.query_params.get('ward_id')
        if ward_id:
            monitors = monitors.filter(room__ward_id=ward_id)
        search = (request.query_params.get('search') or '').strip().lower()
        if search:
            from django.db.models import Q
            monitors = monitors.filter(
                Q(patient_name__icontains=search) | Q(bed_label__icontains=search) | Q(patient_identifier__icontains=search)
            )
        result = []
        critical_count = warning_count = 0
        for pm in monitors:
            last = VitalReading.objects.filter(patient_monitor=pm).order_by('-timestamp').first()
            unack_alarms = Alarm.objects.filter(patient_monitor=pm, acknowledged_at__isnull=True).count()
            status = _vital_status_from_reading(last)
            if unack_alarms > 0:
                status = 'critical'
            if status == 'critical':
                critical_count += 1
            elif status == 'warning':
                warning_count += 1
            ews_score, ews_level = _ews_from_reading(last)
            result.append({
                'id': pm.id,
                'patient_name': pm.patient_name or pm.bed_label or f'Bed {pm.id}',
                'bed_label': pm.bed_label,
                'room_name': pm.room.name,
                'ward_name': pm.room.ward.name if pm.room.ward else None,
                'device_serial': pm.device.serial_number,
                'device_status': effective_device_status(pm.device),
                'bed_status': pm.bed_status,
                'assigned_to': pm.assigned_to_id,
                'assigned_to_name': pm.assigned_to.name if pm.assigned_to else None,
                'last_vital': VitalReadingSerializer(last).data if last else None,
                'unack_alarm_count': unack_alarms,
                'vital_status': status,
                'ews_score': ews_score,
                'ews_level': ews_level,
                'fall_risk': _fall_risk_from_pm(pm, last),
                'pressure_risk': _pressure_risk_from_pm(pm, last),
            })
        status_filter = request.query_params.get('status')
        if status_filter in ('critical', 'warning', 'normal'):
            result = [r for r in result if r['vital_status'] == status_filter]
        summary = {
            'total_beds': len(result),
            'critical_count': sum(1 for r in result if r['vital_status'] == 'critical'),
            'warning_count': sum(1 for r in result if r['vital_status'] == 'warning'),
        }
        return Response({'success': True, 'data': result, 'summary': summary})


class AuditLogList(APIView):
    """GET /api/monitoring/audit-log/ – event log (patient_monitor_id, limit)."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        pm_id = request.query_params.get('patient_monitor_id')
        if pm_id:
            qs = MonitoringAuditLog.objects.filter(patient_monitor_id=pm_id).select_related('user').order_by('-created_at')[:100]
        else:
            qs = MonitoringAuditLog.objects.all().select_related('patient_monitor', 'user').order_by('-created_at')[:200]
        data = MonitoringAuditLogSerializer(qs, many=True).data
        return Response({'success': True, 'data': data})


class MonitoringNoteListCreate(APIView):
    """GET/POST /api/monitoring/notes/ – bemor eslatmalari."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        pm_id = request.query_params.get('patient_monitor_id')
        if not pm_id:
            return Response({'success': True, 'data': []})
        notes = MonitoringNote.objects.filter(patient_monitor_id=pm_id).select_related('created_by').order_by('-created_at')[:50]
        return Response({'success': True, 'data': MonitoringNoteSerializer(notes, many=True).data})

    def post(self, request):
        pm_id = request.data.get('patient_monitor_id')
        note_text = (request.data.get('note') or '').strip()
        if not pm_id or not note_text:
            return Response({'success': False, 'error': 'patient_monitor_id and note required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            pm = PatientMonitor.objects.get(pk=pm_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'Patient monitor not found'}, status=status.HTTP_404_NOT_FOUND)
        note = MonitoringNote.objects.create(
            patient_monitor=pm,
            note=note_text,
            created_by=request.user,
        )
        MonitoringAuditLog.objects.create(patient_monitor=pm, action='note_added', user=request.user, details={'note_id': note.id})
        return Response(MonitoringNoteSerializer(note).data, status=status.HTTP_201_CREATED)


def _vitals_export_rows(qs):
    """Yield (timestamp, heart_rate, spo2, nibp_systolic, nibp_diastolic, respiration_rate, temperature) for export."""
    for r in qs:
        yield (
            r.timestamp.isoformat() if r.timestamp else '',
            r.heart_rate,
            r.spo2,
            r.nibp_systolic,
            r.nibp_diastolic,
            r.respiration_rate,
            float(r.temperature) if r.temperature is not None else '',
        )


class VitalsExport(APIView):
    """GET /api/monitoring/vitals/export/?patient_monitor_id=&from=&to=&format=csv|excel|pdf"""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        pm_id = request.query_params.get('patient_monitor_id')
        from_ts = request.query_params.get('from')
        to_ts = request.query_params.get('to')
        fmt = (request.query_params.get('format') or 'csv').lower()
        if not pm_id:
            return Response({'success': False, 'error': 'patient_monitor_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            pm = PatientMonitor.objects.get(pk=pm_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        qs = VitalReading.objects.filter(patient_monitor=pm).order_by('timestamp')
        if from_ts:
            qs = qs.filter(timestamp__gte=from_ts)
        if to_ts:
            qs = qs.filter(timestamp__lte=to_ts)
        qs = qs[:5000]
        rows = list(_vitals_export_rows(qs))
        headers = ['timestamp', 'heart_rate', 'spo2', 'nibp_systolic', 'nibp_diastolic', 'respiration_rate', 'temperature']
        filename_base = f"vitals_{pm.id}_{(pm.patient_name or pm.bed_label or 'patient').replace(' ', '_')}"

        if fmt == 'csv':
            import csv
            from io import StringIO
            buf = StringIO()
            w = csv.writer(buf)
            w.writerow(headers)
            w.writerows(rows)
            from django.http import HttpResponse
            resp = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
            resp['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
            return resp

        if fmt == 'excel' or fmt == 'xlsx':
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                from django.http import HttpResponse
                wb = Workbook()
                ws = wb.active
                ws.title = 'Vitals'
                for col, h in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=h)
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, val in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 14
                from io import BytesIO
                buf = BytesIO()
                wb.save(buf)
                resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                resp['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
                return resp
            except ImportError:
                return Response({'success': False, 'error': 'Excel export requires openpyxl'}, status=status.HTTP_501_NOT_IMPLEMENTED)

        if fmt == 'pdf':
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib.units import mm
                from io import BytesIO
                from django.http import HttpResponse
                buf = BytesIO()
                doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
                elements = []
                styles = getSampleStyleSheet()
                title = Paragraph(f"Vital signs report — {pm.patient_name or pm.bed_label or f'Patient #{pm.id}'}", styles['Title'])
                elements.append(title)
                elements.append(Spacer(1, 12))
                table_data = [headers] + [[str(c) for c in row] for row in rows]
                t = Table(table_data, colWidths=[50, 22, 22, 28, 28, 28, 28])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
                ]))
                elements.append(t)
                doc.build(elements)
                resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
                resp['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
                return resp
            except ImportError:
                return Response({'success': False, 'error': 'PDF export requires reportlab'}, status=status.HTTP_501_NOT_IMPLEMENTED)

        return Response({'success': False, 'error': 'format must be csv, excel or pdf'}, status=status.HTTP_400_BAD_REQUEST)


class MonitoringStaffList(APIView):
    """GET /api/monitoring/staff/ – list users that can be assigned to beds (monitoring, doctor, staff)."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        User = get_user_model()
        users = User.objects.filter(
            is_active=True,
            role__in=('monitoring', 'doctor', 'staff'),
        ).order_by('name').values('id', 'name', 'phone', 'role')
        return Response({'success': True, 'data': list(users)})


class VitalsCompare(APIView):
    """GET /api/monitoring/vitals/compare/?patient_monitor_id=&range1_from=&range1_to=&range2_from=&range2_to= – two time ranges for trend comparison."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        pm_id = request.query_params.get('patient_monitor_id')
        r1_from = request.query_params.get('range1_from')
        r1_to = request.query_params.get('range1_to')
        r2_from = request.query_params.get('range2_from')
        r2_to = request.query_params.get('range2_to')
        if not pm_id or not all([r1_from, r1_to, r2_from, r2_to]):
            return Response(
                {'success': False, 'error': 'patient_monitor_id, range1_from, range1_to, range2_from, range2_to required'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            pm = PatientMonitor.objects.get(pk=pm_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        base_qs = VitalReading.objects.filter(patient_monitor=pm).order_by('timestamp')
        range1 = list(
            base_qs.filter(timestamp__gte=r1_from, timestamp__lte=r1_to)[:2000].values(
                'timestamp', 'heart_rate', 'spo2', 'nibp_systolic', 'nibp_diastolic', 'respiration_rate', 'temperature'
            )
        )
        range2 = list(
            base_qs.filter(timestamp__gte=r2_from, timestamp__lte=r2_to)[:2000].values(
                'timestamp', 'heart_rate', 'spo2', 'nibp_systolic', 'nibp_diastolic', 'respiration_rate', 'temperature'
            )
        )
        for r in range1:
            if r.get('timestamp'):
                r['timestamp'] = r['timestamp'].isoformat()
        for r in range2:
            if r.get('timestamp'):
                r['timestamp'] = r['timestamp'].isoformat()
        return Response({'success': True, 'data': {'range1': range1, 'range2': range2}})


# --- Quick action, shift report, ward round PDF, medications, lab, stats, family view, rapid response ---

class QuickActionView(APIView):
    """POST /api/monitoring/quick-action/ – Shifokor chaqirish, Lab so'rash, Rapid response (audit log)."""
    permission_classes = [MonitoringPermission]

    def post(self, request):
        pm_id = request.data.get('patient_monitor_id')
        action_type = request.data.get('action_type') or request.data.get('action')  # call_doctor, request_lab, call_family, rapid_response
        note = (request.data.get('note') or '').strip()
        if not pm_id:
            return Response({'success': False, 'error': 'patient_monitor_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            pm = PatientMonitor.objects.get(pk=pm_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        MonitoringAuditLog.objects.create(
            patient_monitor=pm,
            action='quick_action',
            user=request.user,
            details={'action_type': action_type or 'unknown', 'note': note},
        )
        return Response({'success': True, 'message': 'Qeyd qilindi'})


class ShiftReportPDF(APIView):
    """GET /api/monitoring/shift-report/?ward_id=&format=pdf – barcha bemorlar uchun shift hisoboti PDF."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        ward_id = request.query_params.get('ward_id')
        monitors = PatientMonitor.objects.filter(is_active=True).select_related('device', 'room', 'room__ward')
        if ward_id:
            monitors = monitors.filter(room__ward_id=ward_id)
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet
            from io import BytesIO
            from django.http import HttpResponse
        except ImportError:
            return Response({'success': False, 'error': 'reportlab required'}, status=status.HTTP_501_NOT_IMPLEMENTED)
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = [Paragraph('Shift handover hisoboti', styles['Title']), Spacer(1, 12)]
        for pm in monitors[:100]:
            last = VitalReading.objects.filter(patient_monitor=pm).order_by('-timestamp').first()
            unack = Alarm.objects.filter(patient_monitor=pm, acknowledged_at__isnull=True).count()
            line = f"<b>{pm.patient_name or pm.bed_label or f'#{pm.id}'}</b> — {pm.room.name}, kravat {pm.bed_label or '-'}. "
            if last:
                line += f"HR {last.heart_rate or '-'}, SpO2 {last.spo2 or '-'}, AQB {last.nibp_systolic or '-'}/{last.nibp_diastolic or '-'}. "
            if unack:
                line += f"Qabul qilinmagan alarm: {unack}. "
            elements.append(Paragraph(line, styles['Normal']))
            elements.append(Spacer(1, 6))
        doc.build(elements)
        resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="shift_report.pdf"'
        return resp


class WardRoundPDF(APIView):
    """GET /api/monitoring/ward-round-pdf/?patient_monitor_id= – bitta bemor uchun bir sahifa PDF."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        pm_id = request.query_params.get('patient_monitor_id')
        if not pm_id:
            return Response({'success': False, 'error': 'patient_monitor_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            pm = PatientMonitor.objects.get(pk=pm_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from io import BytesIO
            from django.http import HttpResponse
        except ImportError:
            return Response({'success': False, 'error': 'reportlab required'}, status=status.HTTP_501_NOT_IMPLEMENTED)
        last = VitalReading.objects.filter(patient_monitor=pm).order_by('-timestamp').first()
        alarms = Alarm.objects.filter(patient_monitor=pm, acknowledged_at__isnull=True)[:10]
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(f"Ward round — {pm.patient_name or pm.bed_label or f'Bemor #{pm.id}'}", styles['Title']),
            Spacer(1, 8),
            Paragraph(f"Palata: {pm.room.name}, Kravat: {pm.bed_label or '-'}", styles['Normal']),
        ]
        if last:
            elements.append(Paragraph(
                f"Vitals: HR {last.heart_rate or '-'} bpm, SpO2 {last.spo2 or '-'} %, AQB {last.nibp_systolic or '-'}/{last.nibp_diastolic or '-'} mmHg, Nafas {last.respiration_rate or '-'}, Temp {last.temperature or '-'}",
                styles['Normal']
            ))
        if alarms:
            elements.append(Paragraph('Aktiv alarmlar: ' + ', '.join(f"{a.param}={a.value}" for a in alarms), styles['Normal']))
        doc.build(elements)
        resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="ward_round_{pm.id}.pdf"'
        return resp


class MedicationListCreate(APIView):
    """GET/POST /api/monitoring/medications/ – dori eslatmalari."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        pm_id = request.query_params.get('patient_monitor_id')
        if not pm_id:
            return Response({'success': True, 'data': []})
        from django.utils import timezone
        now = timezone.now()
        qs = MonitoringMedication.objects.filter(patient_monitor_id=pm_id).order_by('scheduled_at')
        data = [
            {'id': m.id, 'patient_monitor': m.patient_monitor_id, 'name': m.name, 'dose': m.dose,
             'scheduled_at': m.scheduled_at.isoformat(), 'given_at': m.given_at.isoformat() if m.given_at else None,
             'created_at': m.created_at.isoformat(), 'is_past': m.scheduled_at < now and not m.given_at}
            for m in qs[:100]
        ]
        return Response({'success': True, 'data': data})

    def post(self, request):
        pm_id = request.data.get('patient_monitor_id')
        name = (request.data.get('name') or '').strip()
        dose = (request.data.get('dose') or '').strip()
        scheduled_at = request.data.get('scheduled_at')
        if not pm_id or not name or not scheduled_at:
            return Response({'success': False, 'error': 'patient_monitor_id, name, scheduled_at required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            pm = PatientMonitor.objects.get(pk=pm_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        try:
            from datetime import datetime
            dt = datetime.fromisoformat(scheduled_at.replace('Z', '+00:00'))
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt)
        except Exception:
            return Response({'success': False, 'error': 'Invalid scheduled_at (use ISO format)'}, status=status.HTTP_400_BAD_REQUEST)
        m = MonitoringMedication.objects.create(
            patient_monitor=pm, name=name, dose=dose, scheduled_at=dt, created_by=request.user
        )
        return Response({'success': True, 'data': {'id': m.id, 'scheduled_at': m.scheduled_at.isoformat()}}, status=status.HTTP_201_CREATED)


class MedicationMarkGiven(APIView):
    """POST /api/monitoring/medications/<id>/mark-given/ – Bajarildi belgilash."""
    permission_classes = [MonitoringPermission]

    def post(self, request, pk):
        try:
            m = MonitoringMedication.objects.get(pk=pk)
        except MonitoringMedication.DoesNotExist:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        m.given_at = timezone.now()
        m.save(update_fields=['given_at'])
        return Response({'success': True})


class LabResultListCreate(APIView):
    """GET/POST /api/monitoring/lab-results/ – lab natijalari."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        pm_id = request.query_params.get('patient_monitor_id')
        if not pm_id:
            return Response({'success': True, 'data': []})
        qs = MonitoringLabResult.objects.filter(patient_monitor_id=pm_id).order_by('-timestamp')[:50]
        data = [{'id': r.id, 'patient_monitor': r.patient_monitor_id, 'param': r.param, 'value': r.value, 'unit': r.unit, 'timestamp': r.timestamp.isoformat()} for r in qs]
        return Response({'success': True, 'data': data})

    def post(self, request):
        pm_id = request.data.get('patient_monitor_id')
        param = (request.data.get('param') or '').strip()
        value = (request.data.get('value') or '').strip()
        unit = (request.data.get('unit') or '').strip()
        ts = request.data.get('timestamp')
        if not pm_id or not param or not value:
            return Response({'success': False, 'error': 'patient_monitor_id, param, value required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            pm = PatientMonitor.objects.get(pk=pm_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        timestamp = timezone.now()
        if ts:
            try:
                from datetime import datetime
                timestamp = datetime.fromisoformat(ts.replace('Z', '+00:00'))
                if timezone.is_naive(timestamp):
                    timestamp = timezone.make_aware(timestamp)
            except Exception:
                pass
        r = MonitoringLabResult.objects.create(patient_monitor=pm, param=param, value=value, unit=unit, timestamp=timestamp)
        return Response({'success': True, 'data': {'id': r.id, 'timestamp': r.timestamp.isoformat()}}, status=status.HTTP_201_CREATED)


class AlarmResponseStats(APIView):
    """GET /api/monitoring/alarm-response-stats/ – o'rtacha javob vaqti (room_id, ward_id optional)."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        qs = Alarm.objects.filter(acknowledged_at__isnull=False).select_related('patient_monitor')
        room_id = request.query_params.get('room_id')
        ward_id = request.query_params.get('ward_id')
        if room_id:
            qs = qs.filter(patient_monitor__room_id=room_id)
        if ward_id:
            qs = qs.filter(patient_monitor__room__ward_id=ward_id)
        total_seconds = 0
        count = 0
        for a in qs[:500]:
            delta = a.acknowledged_at - a.created_at
            total_seconds += delta.total_seconds()
            count += 1
        avg_seconds = total_seconds / count if count else None
        return Response({'success': True, 'data': {'avg_response_seconds': avg_seconds, 'sample_count': count}})


class AlarmHeatmap(APIView):
    """GET /api/monitoring/alarm-heatmap/ – xona va soat bo'yicha alarmlar soni."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        from django.db.models import Count
        from django.db.models.functions import TruncHour
        qs = Alarm.objects.annotate(hour=TruncHour('created_at')).values(
            'patient_monitor__room__name', 'hour'
        ).annotate(cnt=Count('id')).order_by('patient_monitor__room__name', 'hour')
        by_room_hour = {}
        for r in qs:
            room_name = r.get('patient_monitor__room__name') or 'N/A'
            hour = r.get('hour')
            if hour:
                hour_key = hour.strftime('%H') if hasattr(hour, 'strftime') else str(hour)
            else:
                hour_key = '?'
            key = (room_name, hour_key)
            by_room_hour[key] = by_room_hour.get(key, 0) + (r.get('cnt') or 0)
        data = [{'room': k[0], 'hour': k[1], 'count': v} for k, v in by_room_hour.items()]
        return Response({'success': True, 'data': data})


class BedForecast(APIView):
    """GET /api/monitoring/bed-forecast/ – kutilayotgan bo'sh kravatlar (bed_status=empty yoki eslatma orqali)."""
    permission_classes = [MonitoringPermission]

    def get(self, request):
        from django.utils import timezone
        from datetime import timedelta
        today = timezone.now().date()
        tomorrow = today + timedelta(days=1)
        empty = PatientMonitor.objects.filter(is_active=True, bed_status='empty').select_related('room')
        occupied = PatientMonitor.objects.filter(is_active=True, bed_status='occupied').select_related('room')
        data = {
            'empty_now': [{'id': pm.id, 'room_name': pm.room.name, 'bed_label': pm.bed_label} for pm in empty[:50]],
            'occupied_count': occupied.count(),
        }
        return Response({'success': True, 'data': data})


def _generate_token():
    import secrets
    return secrets.token_urlsafe(32)


class FamilyViewTokenCreate(APIView):
    """POST /api/monitoring/family-view-token/ – oilaviy ko'rinish uchun token yaratish."""
    permission_classes = [MonitoringPermission]

    def post(self, request):
        pm_id = request.data.get('patient_monitor_id')
        hours = int(request.data.get('hours') or 24)
        if not pm_id:
            return Response({'success': False, 'error': 'patient_monitor_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            pm = PatientMonitor.objects.get(pk=pm_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        from datetime import timedelta
        expires = timezone.now() + timedelta(hours=hours)
        token = _generate_token()
        FamilyViewToken.objects.create(patient_monitor=pm, token=token, expires_at=expires)
        link = request.build_absolute_uri(f'/api/monitoring/family-view/?token={token}')
        return Response({'success': True, 'data': {'token': token, 'link': link, 'expires_at': expires.isoformat()}})


class FamilyViewRead(APIView):
    """GET /api/monitoring/family-view/?token= – token orqali faqat o'qish (auth yo'q)."""
    permission_classes = []  # no auth
    authentication_classes = []

    def get(self, request):
        token = request.query_params.get('token')
        if not token:
            return Response({'success': False, 'error': 'token required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            fv = FamilyViewToken.objects.select_related('patient_monitor').get(token=token)
        except FamilyViewToken.DoesNotExist:
            return Response({'success': False, 'error': 'Invalid or expired token'}, status=status.HTTP_404_NOT_FOUND)
        if timezone.now() > fv.expires_at:
            return Response({'success': False, 'error': 'Token expired'}, status=status.HTTP_403_FORBIDDEN)
        pm = fv.patient_monitor
        last = VitalReading.objects.filter(patient_monitor=pm).order_by('-timestamp').first()
        status_text = 'Stabil'
        if last:
            if (last.spo2 is not None and last.spo2 < 90) or (last.heart_rate is not None and (last.heart_rate > 130 or last.heart_rate < 40)):
                status_text = 'Kuzatilmoqda'
        data = {
            'patient_name': pm.patient_name or pm.bed_label or f'Bemor #{pm.id}',
            'room_name': pm.room.name,
            'status': status_text,
            'heart_rate': last.heart_rate if last else None,
            'spo2': last.spo2 if last else None,
            'updated': last.timestamp.isoformat() if last and last.timestamp else None,
        }
        return Response({'success': True, 'data': data})


class RapidResponseView(APIView):
    """POST /api/monitoring/rapid-response/ – Code blue / Rapid response tugmasi (audit log)."""
    permission_classes = [MonitoringPermission]

    def post(self, request):
        pm_id = request.data.get('patient_monitor_id')
        note = (request.data.get('note') or '').strip()
        if not pm_id:
            return Response({'success': False, 'error': 'patient_monitor_id required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            pm = PatientMonitor.objects.get(pk=pm_id, is_active=True)
        except PatientMonitor.DoesNotExist:
            return Response({'success': False, 'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        MonitoringAuditLog.objects.create(
            patient_monitor=pm,
            action='rapid_response',
            user=request.user,
            details={'note': note},
        )
        return Response({'success': True, 'message': 'Rapid response qayd qilindi'})
